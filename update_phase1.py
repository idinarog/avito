#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скрипт для применения изменений Этапа 1:
- Проверка работоспособности
- Установка зависимостей
- Создание резервной копии
- Обновление requirements.txt
"""

import os
import sys
import subprocess
import shutil
from pathlib import Path
from datetime import datetime


class Phase1Updater:
    """Класс для применения изменений Этапа 1"""
    
    def __init__(self):
        self.project_root = Path(__file__).parent
        self.backup_dir = self.project_root / f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.errors = []
        self.steps = []
        
        # Цвета для вывода
        self.GREEN = '\033[92m'
        self.YELLOW = '\033[93m'
        self.RED = '\033[91m'
        self.BLUE = '\033[94m'
        self.RESET = '\033[0m'
        self.BOLD = '\033[1m'
    
    def print_header(self, text):
        """Вывод заголовка"""
        print(f"\n{self.BOLD}{self.BLUE}{'='*60}{self.RESET}")
        print(f"{self.BOLD}{self.BLUE}{text:^60}{self.RESET}")
        print(f"{self.BOLD}{self.BLUE}{'='*60}{self.RESET}\n")
    
    def print_success(self, text):
        print(f"{self.GREEN}✅ {text}{self.RESET}")
        self.steps.append(f"✅ {text}")
    
    def print_error(self, text):
        print(f"{self.RED}❌ {text}{self.RESET}")
        self.errors.append(f"❌ {text}")
    
    def print_info(self, text):
        print(f"{self.BLUE}ℹ️ {text}{self.RESET}")
    
    def print_warning(self, text):
        print(f"{self.YELLOW}⚠️ {text}{self.RESET}")
    
    def check_python_version(self) -> bool:
        """Проверка версии Python"""
        self.print_header("1. ПРОВЕРКА ВЕРСИИ PYTHON")
        
        version = sys.version_info
        print(f"Текущая версия: Python {version.major}.{version.minor}.{version.micro}")
        
        if version.major < 3 or (version.major == 3 and version.minor < 8):
            self.print_error(f"Требуется Python 3.8+, у вас {version.major}.{version.minor}")
            return False
        
        self.print_success(f"Python {version.major}.{version.minor} — OK")
        return True
    
    def check_venv(self) -> bool:
        """Проверка виртуального окружения"""
        self.print_header("2. ПРОВЕРКА ВИРТУАЛЬНОГО ОКРУЖЕНИЯ")
        
        in_venv = hasattr(sys, 'real_prefix') or (hasattr(sys, 'base_prefix') and sys.base_prefix != sys.prefix)
        
        if not in_venv:
            self.print_warning("Вы не в виртуальном окружении!")
            response = input("Продолжить без виртуального окружения? (y/n): ")
            if response.lower() != 'y':
                self.print_error("Установка отменена")
                return False
        
        self.print_success("Виртуальное окружение активно" if in_venv else "Продолжаем без виртуального окружения")
        return True
    
    def install_dependencies(self) -> bool:
        """Установка новых зависимостей"""
        self.print_header("3. УСТАНОВКА НОВЫХ ЗАВИСИМОСТЕЙ")
        
        new_packages = [
            "pandas",
            "openpyxl",
            "apscheduler",
        ]
        
        print("Будут установлены:")
        for pkg in new_packages:
            print(f"  📦 {pkg}")
        
        response = input("\nУстановить новые зависимости? (y/n): ")
        if response.lower() != 'y':
            self.print_warning("Установка зависимостей пропущена")
            return True
        
        success = True
        for pkg in new_packages:
            try:
                print(f"Установка {pkg}...")
                subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])
                self.print_success(f"Установлен {pkg}")
            except subprocess.CalledProcessError as e:
                self.print_error(f"Ошибка установки {pkg}: {e}")
                success = False
        
        return success
    
    def update_requirements(self) -> bool:
        """Обновление requirements.txt"""
        self.print_header("4. ОБНОВЛЕНИЕ REQUIREMENTS.TXT")
        
        req_path = self.project_root / "requirements.txt"
        
        # Читаем существующий файл
        existing = []
        if req_path.exists():
            with open(req_path, 'r', encoding='utf-8') as f:
                existing = [line.strip() for line in f if line.strip()]
        
        # Добавляем новые зависимости
        new_deps = [
            "pandas>=1.5.0",
            "openpyxl>=3.0.0",
            "apscheduler>=3.9.0",
        ]
        
        # Объединяем, удаляем дубликаты
        all_deps = list(set(existing + new_deps))
        all_deps.sort()
        
        with open(req_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(all_deps))
            f.write('\n')
        
        self.print_success(f"Обновлен {req_path}")
        self.print_info(f"Всего зависимостей: {len(all_deps)}")
        return True
    
    def create_backup(self) -> bool:
        """Создание резервной копии"""
        self.print_header("5. СОЗДАНИЕ РЕЗЕРВНОЙ КОПИИ")
        
        print(f"Будет создана резервная копия в: {self.backup_dir}")
        
        response = input("Создать резервную копию? (y/n): ")
        if response.lower() != 'y':
            self.print_warning("Создание резервной копии пропущено")
            return True
        
        try:
            # Создаем папку для бэкапа
            self.backup_dir.mkdir(parents=True, exist_ok=True)
            
            # Копируем важные файлы и папки
            items_to_backup = [
                "main.py",
                "core",
                "ui",
                "api",
                "database",
                "services",
                "utils",
                "requirements.txt",
                ".env",
            ]
            
            for item in items_to_backup:
                src = self.project_root / item
                dst = self.backup_dir / item
                
                if src.exists():
                    if src.is_dir():
                        shutil.copytree(src, dst, ignore=shutil.ignore_patterns('__pycache__', '*.pyc', '.DS_Store'))
                    else:
                        shutil.copy2(src, dst)
                    print(f"  📁 {item} → {dst}")
            
            self.print_success(f"Резервная копия создана: {self.backup_dir}")
            return True
            
        except Exception as e:
            self.print_error(f"Ошибка создания резервной копии: {e}")
            return False
    
    def check_project_structure(self) -> bool:
        """Проверка структуры проекта"""
        self.print_header("6. ПРОВЕРКА СТРУКТУРЫ ПРОЕКТА")
        
        required_items = [
            "main.py",
            "core/app.py",
            "ui/main_window.py",
            "database/models.py",
            "api/avito_api.py",
            ".env",
        ]
        
        all_ok = True
        for item in required_items:
            path = self.project_root / item
            if path.exists():
                print(f"  ✅ {item}")
            else:
                print(f"  ❌ {item} (не найден)")
                all_ok = False
        
        if all_ok:
            self.print_success("Структура проекта OK")
        else:
            self.print_warning("Некоторые файлы отсутствуют, но это не критично")
        
        return True
    
    def check_imports(self) -> bool:
        """Проверка импортов"""
        self.print_header("7. ПРОВЕРКА ИМПОРТОВ")
        
        try:
            import pandas
            import openpyxl
            from apscheduler.schedulers.qt import QtScheduler
            
            self.print_success("✅ pandas импортируется")
            self.print_success("✅ openpyxl импортируется")
            self.print_success("✅ apscheduler импортируется")
            return True
            
        except ImportError as e:
            self.print_error(f"Ошибка импорта: {e}")
            return False
    
    def create_test_script(self) -> bool:
        """Создание тестового скрипта для проверки"""
        self.print_header("8. СОЗДАНИЕ ТЕСТОВОГО СКРИПТА")
        
        test_script = self.project_root / "test_phase1.py"
        
        content = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Тестовый скрипт для проверки Этапа 1
Проверяет установку зависимостей и импорты
"""

import sys
import pandas as pd
import openpyxl
from apscheduler.schedulers.qt import QtScheduler

def test_pandas():
    """Проверка pandas"""
    df = pd.DataFrame({'col1': [1, 2, 3], 'col2': ['a', 'b', 'c']})
    print(f"✅ pandas работает: {df.shape}")
    return True

def test_openpyxl():
    """Проверка openpyxl"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    print(f"✅ openpyxl работает: {ws['A1'].value}")
    return True

def test_apscheduler():
    """Проверка apscheduler"""
    from apscheduler.triggers.interval import IntervalTrigger
    trigger = IntervalTrigger(minutes=5)
    print(f"✅ apscheduler работает: {trigger}")
    return True

def main():
    print("🧪 Тестирование Этапа 1")
    print("=" * 40)
    
    tests = [
        test_pandas,
        test_openpyxl,
        test_apscheduler,
    ]
    
    all_ok = True
    for test in tests:
        try:
            if test():
                print("✅")
            else:
                print("❌")
                all_ok = False
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            all_ok = False
    
    if all_ok:
        print("\\n" + "=" * 40)
        print("✅ Все тесты пройдены! Можно переходить к Этапу 2.")
    else:
        print("\\n" + "=" * 40)
        print("❌ Некоторые тесты не пройдены. Проверьте установку зависимостей.")

if __name__ == "__main__":
    main()
'''
        
        with open(test_script, 'w', encoding='utf-8') as f:
            f.write(content)
        
        # Делаем исполняемым
        os.chmod(test_script, 0o755)
        
        self.print_success(f"Создан тестовый скрипт: {test_script}")
        return True
    
    def show_summary(self):
        """Показывает итоговый отчет"""
        self.print_header("📊 ИТОГОВЫЙ ОТЧЕТ")
        
        print(f"\n{self.BOLD}Выполненные шаги:{self.RESET}")
        for step in self.steps:
            print(f"  {step}")
        
        if self.errors:
            print(f"\n{self.RED}Ошибки:{self.RESET}")
            for error in self.errors:
                print(f"  {error}")
        
        print(f"\n{self.BOLD}{self.GREEN}✅ Этап 1 завершен!{self.RESET}")
        print(f"\n{self.BOLD}📋 Дальнейшие действия:{self.RESET}")
        print(f"  1. Запустите тестовый скрипт: {self.BOLD}python test_phase1.py{self.RESET}")
        print(f"  2. Если всё OK, переходите к Этапу 2")
        print(f"  3. При возникновении проблем: восстановитесь из бэкапа")
        print(f"     {self.backup_dir}")
    
    def run(self):
        """Запуск всех проверок и обновлений"""
        print(self.BOLD + """
    ╔═══════════════════════════════════════════════════════════╗
    ║         🔧 ЭТАП 1: АУДИТ И ПОДГОТОВКА                  ║
    ║         Avito Commander 1.0                            ║
    ╚═══════════════════════════════════════════════════════════╝
    """ + self.RESET)
        
        # Проверка Python
        if not self.check_python_version():
            return
        
        # Проверка виртуального окружения
        if not self.check_venv():
            return
        
        # Установка зависимостей
        if not self.install_dependencies():
            self.print_warning("Проблемы с установкой зависимостей, но продолжаем...")
        
        # Обновление requirements.txt
        if not self.update_requirements():
            self.print_warning("Не удалось обновить requirements.txt")
        
        # Создание резервной копии
        if not self.create_backup():
            self.print_warning("Проблемы с созданием резервной копии")
        
        # Проверка структуры
        self.check_project_structure()
        
        # Проверка импортов
        self.check_imports()
        
        # Создание тестового скрипта
        self.create_test_script()
        
        # Итоговый отчет
        self.show_summary()


def main():
    updater = Phase1Updater()
    updater.run()


if __name__ == "__main__":
    main()