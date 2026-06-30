#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Тестовый скрипт для проверки Этапа 1
Проверяет установку зависимостей и импорты
"""

import sys
import pandas as pd
import openpyxl
from apscheduler.schedulers.qt import QtScheduler

def test_pandas():
    """Проверка pandas"""
    df = pd.DataFrame({'col1': [1, 2, 3], 'col2': ['a', 'b', 'c']})
    print(f"✅ pandas работает: {df.shape}")
    return True

def test_openpyxl():
    """Проверка openpyxl"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    print(f"✅ openpyxl работает: {ws['A1'].value}")
    return True

def test_apscheduler():
    """Проверка apscheduler"""
    from apscheduler.triggers.interval import IntervalTrigger
    trigger = IntervalTrigger(minutes=5)
    print(f"✅ apscheduler работает: {trigger}")
    return True

def main():
    print("🧪 Тестирование Этапа 1")
    print("=" * 40)
    
    tests = [
        test_pandas,
        test_openpyxl,
        test_apscheduler,
    ]
    
    all_ok = True
    for test in tests:
        try:
            if test():
                print("✅")
            else:
                print("❌")
                all_ok = False
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            all_ok = False
    
    if all_ok:
        print("\n" + "=" * 40)
        print("✅ Все тесты пройдены! Можно переходить к Этапу 2.")
    else:
        print("\n" + "=" * 40)
        print("❌ Некоторые тесты не пройдены. Проверьте установку зависимостей.")

if __name__ == "__main__":
    main()
